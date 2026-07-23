import os
import re
import csv
import json
import sqlite3
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

def get_db_path(company_id: str = None) -> str:
    """
    Per-tenant SQLite database path: <knowledge_base>/db/<company_id>_tables.db.
    Different companies get physically separate .db files so the SQL router can
    never read another tenant's tables. Falls back to the default company.
    """
    from backend.src import config
    cid = config.normalize_company_id(company_id or config.DEFAULT_COMPANY_ID)
    db_dir = os.path.join(config.DOCUMENTS_DIR, "db")
    os.makedirs(db_dir, exist_ok=True)
    return os.path.join(db_dir, f"{cid}_tables.db")

def sanitize_name(name: str) -> str:
    """Sanitize filename, sheet name or table ID to be a safe SQL table name."""
    clean = re.sub(r'[^\w]', '_', name)
    clean = re.sub(r'_+', '_', clean)
    clean = clean.lower().strip("_")
    if clean and clean[0].isdigit():
        clean = "t_" + clean
    return clean

def coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert columns to numeric where possible, keeping non-numeric columns as-is.
    Replaces df.apply(pd.to_numeric, errors='ignore'): 'ignore' was removed in
    pandas 3.x and raises "ValueError: invalid error value specified" — which
    broke indexing on fresh installs (e.g. Railway) while older local pandas
    only warned. This per-column try/except works on every pandas version.
    """
    def _maybe_numeric(col):
        try:
            return pd.to_numeric(col)
        except (ValueError, TypeError):
            return col
    return df.apply(_maybe_numeric)


def to_json_serializable(val):
    """Convert numpy or pandas types to JSON-serializable native Python types."""
    import datetime
    if pd.isna(val):
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if hasattr(val, "item"):
        try:
            return val.item()
        except Exception:
            pass
    return val

# ── Internal Canonical Model (ICM) Dataclasses & Helpers ──
from dataclasses import dataclass, field
from enum import Enum

class RegionType(str, Enum):
    DATA_TABLE = "data_table"
    KEY_VALUE = "key_value"
    NOTES = "notes"
    DASHBOARD = "dashboard"
    UNKNOWN = "unknown"

@dataclass
class ICMCell:
    raw_val: Optional[object] = None
    display_val: str = ""
    numeric_val: Optional[float] = None
    date_val: Optional[str] = None
    bool_val: Optional[bool] = None

@dataclass
class SheetCapabilities:
    has_tables: bool = True
    supports_sql: bool = True
    supports_cell_search: bool = True
    has_notes: bool = False

@dataclass
class ICMRegion:
    region_id: str
    region_type: RegionType
    table_title: str
    df: pd.DataFrame
    headers: List[str]
    bounds: Tuple[int, int, int, int]  # (min_r, max_r, min_c, max_c)
    confidence: float = 1.0
    confidence_reason: str = "Standard grid component"
    metadata: Dict = field(default_factory=dict)

@dataclass
class ICMSheet:
    sheet_name: str
    regions: List[ICMRegion] = field(default_factory=list)
    capabilities: SheetCapabilities = field(default_factory=SheetCapabilities)

@dataclass
class ICMWorkbook:
    filename: str
    sheets: List[ICMSheet] = field(default_factory=list)
    total_rows: int = 0
    total_cols: int = 0

def extract_cell_representations(val: object) -> ICMCell:
    """Extract multi-type cell representation (raw, display, numeric, date, bool)."""
    if val is None or pd.isna(val):
        return ICMCell(raw_val=None, display_val="", numeric_val=None, date_val=None, bool_val=None)
    
    raw_val = val
    display_val = str(val).strip()
    
    # 1. Check boolean
    bool_val = None
    if isinstance(val, bool):
        bool_val = val
    elif display_val.lower() in ("true", "yes", "y"):
        bool_val = True
    elif display_val.lower() in ("false", "no", "n"):
        bool_val = False

    # 2. Check numeric
    numeric_val = None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        numeric_val = float(val)
    else:
        # Clean currency/comma strings like "$1,250.50" or " 1250 "
        clean_num_str = re.sub(r'[^\d.-]', '', display_val)
        if clean_num_str and clean_num_str not in ("-", ".", "-."):
            try:
                numeric_val = float(clean_num_str)
            except ValueError:
                numeric_val = None

    # 3. Check date
    date_val = None
    if isinstance(val, (datetime, pd.Timestamp)):
        date_val = val.isoformat()
    elif isinstance(val, str):
        # Try simple ISO or standard date formats
        if re.match(r'^\d{4}-\d{2}-\d{2}', display_val):
            date_val = display_val

    return ICMCell(
        raw_val=raw_val,
        display_val=display_val,
        numeric_val=numeric_val,
        date_val=date_val,
        bool_val=bool_val
    )

import hashlib

def compute_semantic_region_hash(region: ICMRegion) -> str:
    """
    Compute a stable MD5 hash based strictly on semantic content
    (RegionType + Headers + Shape + Coordinates + Sample values).
    Serializer formatting changes will NOT alter this hash.
    """
    hasher = hashlib.md5()
    hasher.update(str(region.region_type.value).encode('utf-8'))
    hasher.update(",".join(region.headers).encode('utf-8'))
    hasher.update(str(region.bounds).encode('utf-8'))
    
    if not region.df.empty:
        shape_str = f"{len(region.df)}x{len(region.df.columns)}"
        hasher.update(shape_str.encode('utf-8'))
        sample_vals = str(region.df.head(10).to_dict(orient='records'))
        hasher.update(sample_vals.encode('utf-8'))
        
    return hasher.hexdigest()

class BaseSerializer:
    """Abstract Serializer interface for converting ICMRegion to text/metadata."""
    def serialize(self, region: ICMRegion) -> str:
        raise NotImplementedError

class TOONSerializer(BaseSerializer):
    """TOON (Object-Oriented Normalized JSON) Serializer."""
    def serialize(self, region: ICMRegion) -> str:
        data_dict = {
            "region_id": region.region_id,
            "type": region.region_type.value,
            "title": region.table_title,
            "headers": region.headers,
            "rows_count": len(region.df),
            "bounds": list(region.bounds),
            "confidence": region.confidence,
            "confidence_reason": region.confidence_reason,
            "preview_rows": region.df.head(5).to_dict(orient="records")
        }
        return json.dumps(data_dict, indent=2, default=to_json_serializable)

class MarkdownSerializer(BaseSerializer):
    """Markdown Serializer for vector embeddings & prompt context."""
    def serialize(self, region: ICMRegion) -> str:
        title_str = f" (Title: {region.table_title})" if region.table_title else ""
        preview_md = region.df.head(5).fillna("").to_markdown(index=False)
        return (
            f"## Region: {region.region_id}{title_str} [{region.region_type.value}]\n"
            f"Total Rows: {len(region.df)} | Columns ({len(region.headers)}): {', '.join(region.headers)}\n\n"
            f"### Sample Preview:\n{preview_md}"
        )

class JSONSerializer(BaseSerializer):
    """Full JSON Serializer."""
    def serialize(self, region: ICMRegion) -> str:
        return region.df.to_json(orient="records", default_handler=to_json_serializable)

def detect_csv_properties(file_path: str) -> Tuple[str, str]:
    """Sniff CSV encoding and delimiter defensively."""
    # 1. Detect Encoding
    encodings = ["utf-8", "latin-1", "utf-16", "cp1252"]
    detected_encoding = "utf-8"
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                f.read(2048)
            detected_encoding = enc
            break
        except Exception:
            continue
            
    # 2. Sniff Delimiter
    detected_delimiter = ","
    try:
        with open(file_path, "r", encoding=detected_encoding, errors="ignore") as f:
            sample = f.read(4096)
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            if dialect.delimiter in (",", ";", "\t", "|"):
                detected_delimiter = dialect.delimiter
    except Exception:
        # Fall back to checking common delimiters in sample
        try:
            with open(file_path, "r", encoding=detected_encoding, errors="ignore") as f:
                sample = f.read(2048)
                counts = {d: sample.count(d) for d in (",", ";", "\t", "|")}
                best = max(counts, key=counts.get)
                if counts[best] > 0:
                    detected_delimiter = best
        except Exception:
            pass
            
    return detected_encoding, detected_delimiter

def process_csv_file(file_path: str) -> Tuple[List[Dict], List[Tuple[str, pd.DataFrame, str]]]:
    """
    Parse a CSV file as a single clean table.
    Generates Tier A (row JSON) and Tier B (markdown window) chunks.
    Also returns the table to be loaded into SQLite (Tier C).
    """
    filename = os.path.basename(file_path)
    logger.info(f"Ingesting CSV file: {filename}")
    
    encoding, delimiter = detect_csv_properties(file_path)
    try:
        df = pd.read_csv(file_path, encoding=encoding, sep=delimiter, on_bad_lines="skip")
    except Exception as e:
        logger.error(f"pd.read_csv failed for {filename}: {e}")
        # Fallback to python csv reader if pandas parser breaks
        rows = []
        with open(file_path, "r", encoding=encoding, errors="ignore") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows.append(row)
        if not rows:
            return [], []
        headers = [h.strip() or f"col_{i}" for i, h in enumerate(rows[0])]
        df = pd.DataFrame(rows[1:], columns=headers)

    if df.empty:
        return [], []

    # Clean headers
    cleaned_headers = []
    seen = {}
    for i, h in enumerate(df.columns):
        h_clean = sanitize_name(str(h))
        if not h_clean or "unnamed" in h_clean.lower():
            h_clean = f"col_{i}"
        if h_clean in seen:
            seen[h_clean] += 1
            cleaned_headers.append(f"{h_clean}_{seen[h_clean]}")
        else:
            seen[h_clean] = 1
            cleaned_headers.append(h_clean)
    df.columns = cleaned_headers

    # Infer real types for SQLite
    df = coerce_numeric_columns(df)
    
    db_table_name = sanitize_name(f"{filename}_default_table")
    chunks = []
    dtypes_dict = {col: str(dtype) for col, dtype in df.dtypes.items()}
    
    # ── Table Summary Chunk for Vector Search ──
    preview_df = df.head(5).fillna("")
    preview_md = preview_df.to_markdown(index=False)
    summary_text = (
        f"## Table: {filename} (SQL Table Name: {db_table_name})\n"
        f"Total Rows: {len(df)}\n"
        f"Columns ({len(df.columns)}): {', '.join(df.columns)}\n\n"
        f"### Sample Rows Preview:\n{preview_md}"
    )
    chunks.append({
        "text": summary_text,
        "metadata": {
            "source": file_path,
            "source_name": filename,
            "source_file": filename,
            "type": "csv",
            "sheet_name": "default",
            "table_id": db_table_name,
            "row_range": [1, len(df)],
            "columns": list(df.columns),
            "column_dtypes": dtypes_dict,
            "chunk_tier": "table_summary",
        }
    })

    # Smart windowing: scale window chunks based on table size.
    # Large files (>500 rows): summary only — SQL handles all lookups.
    # Medium files (51-500 rows): summary + up to 5 sampled windows for context.
    # Small files (<=50 rows): summary + all row windows (at most ceil(50/15)=4 extra chunks).
    nrows = len(df)
    if nrows <= 50:
        window_size = 15
        windows_to_emit = list(range(0, nrows, window_size))
    elif nrows <= 500:
        window_size = max(nrows // 5, 15)
        windows_to_emit = list(range(0, nrows, window_size))[:5]
    else:
        windows_to_emit = []  # Summary chunk only — SQL router answers specific queries

    for i in windows_to_emit:
        window_df = df.iloc[i:i+window_size].fillna("")
        window_df = window_df.copy()
        for col in window_df.columns:
            if window_df[col].dtype == object:
                window_df[col] = window_df[col].astype(str).str.replace('\n', ' ', regex=False).str.replace('\r', '', regex=False)
        md_table = window_df.to_markdown(index=False)
        chunk_text = f"## File: {filename} (Rows {i+1}-{min(i+window_size, nrows)})\n\n{md_table}"
        chunks.append({
            "text": chunk_text,
            "metadata": {
                "source": file_path,
                "source_name": filename,
                "source_file": filename,
                "type": "csv",
                "sheet_name": "default",
                "table_id": db_table_name,
                "row_range": [i + 1, min(i + window_size, nrows)],
                "columns": list(df.columns),
                "column_dtypes": dtypes_dict,
                "chunk_tier": "markdown_window"
            }
        })

    # Per-file safety cap (same policy as Excel — see process_excel_file).
    MAX_CHUNKS_PER_FILE = 100
    if len(chunks) > MAX_CHUNKS_PER_FILE:
        logger.warning(
            f"⚠️  CSV {filename}: chunk count {len(chunks)} exceeds cap {MAX_CHUNKS_PER_FILE}. "
            f"Keeping first {MAX_CHUNKS_PER_FILE}."
        )
        summaries = [c for c in chunks if c["metadata"].get("chunk_tier") == "table_summary"]
        windows   = [c for c in chunks if c["metadata"].get("chunk_tier") != "table_summary"]
        chunks = (summaries + windows)[:MAX_CHUNKS_PER_FILE]

    logger.info(f"CSV {filename}: {nrows} rows → {len(chunks)} vector chunk(s) (SQL table also stored).")
    return chunks, [(db_table_name, df, "")]

def is_key_value_block(df: pd.DataFrame) -> bool:
    """Detect if a dataframe is structured as a key-value block (e.g. metadata list)."""
    non_empty_cols = [col for col in df.columns if not df[col].isna().all()]
    if len(non_empty_cols) != 2:
        return False
    col0_vals = df[non_empty_cols[0]].dropna().astype(str).str.strip()
    if len(col0_vals) < 2:
        return False
    
    colon_ends = sum(1 for v in col0_vals if v.endswith(":") or v.lower() in (
        "account", "sales team", "customer success manager", "client spoc", "scope",
        "contract signed on", "contract tenure", "renewal date", "launch date",
        "contract value", "license", "database", "implementation", "integration",
        "support", "success fee", "sms cost", "customization"
    ))
    ratio = colon_ends / len(col0_vals)
    return ratio >= 0.3

def detect_data_start_row(sheet_val, min_r, max_r, filtered_cols_indices, grid_val) -> int:
    """
    Scans row by row from top of the component to detect where table header ends and data begins.
    Uses cell styling (bold/fills) and presence of numeric/formula values as signals.
    """
    if len(filtered_cols_indices) <= 2:
        return 1
        
    num_rows = max_r - min_r + 1
    if num_rows <= 1:
        return 0
        
    for idx, r_idx in enumerate(range(min_r, max_r + 1)):
        row_vals = [grid_val[r_idx][c_idx] for c_idx in filtered_cols_indices]
        non_empty_vals = [v for v in row_vals if v is not None and str(v).strip() != ""]
        if not non_empty_vals:
            continue
            
        # 1. Check for numeric values in row (excluding common year numbers)
        has_numeric = False
        for v in non_empty_vals:
            if isinstance(v, (int, float)):
                if v > 2100 or v < 0 or (v > 20 and v != 2024 and v != 2025 and v != 2026 and v != 2027):
                    has_numeric = True
                    break
            elif isinstance(v, str):
                clean_v = v.replace(",", "").replace("$", "").replace("%", "").strip()
                if clean_v.isdigit():
                    val_int = int(clean_v)
                    if val_int > 2100 or val_int < 0 or (val_int > 20 and val_int != 2024 and val_int != 2025 and val_int != 2026 and val_int != 2027):
                        has_numeric = True
                        break
                else:
                    try:
                        float(clean_v)
                        has_numeric = True
                        break
                    except ValueError:
                        pass
        
        if has_numeric:
            return idx
            
        # 2. Check styling (if a row has low bold/highlight ratio, it is data)
        bold_count = 0
        for c_idx in filtered_cols_indices:
            cell = sheet_val.cell(row=r_idx + 1, column=c_idx + 1)
            is_bold = False
            if cell.font and cell.font.bold:
                is_bold = True
            elif cell.fill and cell.fill.start_color and cell.fill.start_color.rgb not in ("00000000", "000000", "FFFFFFFF", "FFFFFFF"):
                is_bold = True
            if is_bold:
                bold_count += 1
                
        bold_ratio = bold_count / len(non_empty_vals) if non_empty_vals else 0
        if bold_ratio < 0.3:
            return idx
            
    return 1

def is_crosstab_table(df: pd.DataFrame) -> Tuple[bool, int]:
    """
    Check if a DataFrame is a 2D crosstab matrix with multi-row headers.
    Returns (is_crosstab, num_header_rows).
    """
    if len(df) < 3 or len(df.columns) < 3:
        return False, 0

    # Inspect top 2 rows: check if row 0 contains repeated category headers (e.g. Years)
    row0_vals = [str(x).strip() for x in df.iloc[0].values if pd.notna(x) and str(x).strip() != ""]
    if len(row0_vals) >= 2:
        unique_row0 = set(row0_vals)
        if len(unique_row0) < len(row0_vals):
            return True, 2
            
    return False, 0

def unpivot_crosstab_table(df: pd.DataFrame, num_header_rows: int = 2) -> pd.DataFrame:
    """
    Unpivot a 2D matrix report with multi-row headers into a flat relational DataFrame.
    Example:
         Row Metric | 2024 Q1 | 2024 Q2 | 2025 Q1
         Sales      | 100     | 150     | 200
    Becomes:
         Metric | Dimension_1 | Dimension_2 | Value
         Sales  | 2024        | Q1          | 100
    """
    if df.empty or len(df) <= num_header_rows:
        return df

    try:
        metric_col_name = str(df.columns[0]).strip() or "Metric"
        header_levels = []
        for h_idx in range(num_header_rows):
            level_vals = [str(val).strip() if pd.notna(val) else "" for val in df.iloc[h_idx].values]
            header_levels.append(level_vals)

        data_df = df.iloc[num_header_rows:].copy()
        unpivoted_rows = []

        for _, row in data_df.iterrows():
            metric_val = row.iloc[0]
            if pd.isna(metric_val) or str(metric_val).strip() == "":
                continue
            
            for col_idx in range(1, len(df.columns)):
                cell_val = row.iloc[col_idx]
                if pd.isna(cell_val) or str(cell_val).strip() == "":
                    continue
                
                dim_1 = header_levels[0][col_idx] if len(header_levels) > 0 else ""
                dim_2 = header_levels[1][col_idx] if len(header_levels) > 1 else ""
                
                unpivoted_rows.append([str(metric_val).strip(), dim_1, dim_2, cell_val])

        cols = [metric_col_name, "Dimension_1", "Dimension_2", "Value"]
        flat_df = pd.DataFrame(unpivoted_rows, columns=cols)
        return coerce_numeric_columns(flat_df)
    except Exception as e:
        logger.warning(f"unpivot_crosstab_table failed: {e}")
        return df

def classify_region_component(df: pd.DataFrame, min_r: int, max_r: int, min_c: int, max_c: int) -> Tuple[RegionType, float, str]:
    """
    Classify a grid component into RegionType with confidence score and reasoning.
    """
    if df.empty:
        return RegionType.UNKNOWN, 0.0, "Empty region"
        
    nrows, ncols = len(df), len(df.columns)
    
    # 1. Key-Value block
    if is_key_value_block(df):
        return RegionType.KEY_VALUE, 0.95, "2-column key-value layout"
        
    # 2. Regular Data Table (2+ rows, 1+ cols with decent density)
    if nrows >= 2 and ncols >= 1:
        non_empty_ratio = (df.notna().sum().sum()) / (nrows * ncols)
        if non_empty_ratio >= 0.3:
            return RegionType.DATA_TABLE, 0.95, "Structured tabular grid with data"
            
    # 3. Footnote / Notes (single-column short text block or small footnote)
    if nrows <= 3 and ncols == 1:
        return RegionType.NOTES, 0.90, "Short text footnote block"
            
    # 4. Unknown Region fallback
    return RegionType.UNKNOWN, 0.50, "Unstructured grid block preserved as raw region"

PARSER_VERSION = "v6"

def is_same_column_signature(cols_a: List[str], cols_b: List[str]) -> bool:
    """Check if two column sets have >= 90% overlap or matching column count."""
    if not cols_a or not cols_b:
        return False
    if len(cols_a) == len(cols_b):
        # If same length, check match ratio
        matches = sum(1 for a, b in zip(cols_a, cols_b) if a == b or a in cols_b)
        return (matches / len(cols_a)) >= 0.8
    set_a, set_b = set(cols_a), set(cols_b)
    intersection = set_a.intersection(set_b)
    union = set_a.union(set_b)
    if not union:
        return False
    return (len(intersection) / len(union)) >= 0.8

def process_excel_file(file_path: str) -> Tuple[List[Dict], List[Tuple[str, pd.DataFrame, str]]]:
    """
    Parse an Excel (.xlsx, .xls) workbook.
    Runs table segmentation to detect multiple side-by-side or stacked tables per sheet.
    Un-merges cells within specific bounds, flattens multi-row headers,
    captures style highlights (bold/red), and filters hidden columns/rows.
    Merges fragments with matching column signatures to avoid over-segmentation.
    Generates Tier A & Tier B chunks and returns Tier C tables for SQLite/Postgres.
    """
    filename = os.path.basename(file_path)
    logger.info(f"Ingesting Excel file: {filename} (Parser {PARSER_VERSION})")
    
    chunks = []
    sqlite_tables = []
    
    try:
        # Load twice: once data-only for display values, once formula-only
        with open(file_path, "rb") as f:
            wb_val = openpyxl.load_workbook(f, data_only=True)
        with open(file_path, "rb") as f:
            wb_form = openpyxl.load_workbook(f, data_only=False)
    except Exception as e:
        logger.error(f"Failed to load workbook for {filename}: {e}")
        return [], []
        
    all_sheet_tables = []
    leftover_chunks = []

    for sheet_name in wb_val.sheetnames:
        sheet_val = wb_val[sheet_name]
        if sheet_name not in wb_form.sheetnames:
            continue
        sheet_form = wb_form[sheet_name]
        
        max_row = sheet_val.max_row
        max_column = sheet_val.max_column
        if max_row <= 1 or max_column < 1:
            continue
            
        # 1. Resolve and copy grid values, copying top-left cell values into merged cells
        grid_val = [[None for _ in range(max_column)] for _ in range(max_row)]
        grid_form = [[None for _ in range(max_column)] for _ in range(max_row)]
        
        for r in range(1, max_row + 1):
            for c in range(1, max_column + 1):
                grid_val[r-1][c-1] = sheet_val.cell(row=r, column=c).value
                grid_form[r-1][c-1] = sheet_form.cell(row=r, column=c).value
                
        # Forward-fill merged cells safely strictly within bounds
        for merged_range in sheet_val.merged_cells.ranges:
            min_col, min_row, max_col, max_row_range = merged_range.bounds
            min_col = min(min_col, max_column)
            min_row = min(min_row, max_row)
            max_col = min(max_col, max_column)
            max_row_range = min(max_row_range, max_row)
            
            top_left_val = grid_val[min_row-1][min_col-1]
            top_left_form = grid_form[min_row-1][min_col-1]
            
            for r in range(min_row, max_row_range + 1):
                for c in range(min_col, max_col + 1):
                    grid_val[r-1][c-1] = top_left_val
                    grid_form[r-1][c-1] = top_left_form
                    
        # 2. Detect title banner rows
        banner_rows = set()
        if max_column > 2:
            for r in range(max_row):
                non_empty_raw = sum(1 for c in range(max_column) if sheet_val.cell(row=r+1, column=c+1).value is not None and str(sheet_val.cell(row=r+1, column=c+1).value).strip() != "")
                non_empty_grid = sum(1 for c in range(max_column) if grid_val[r][c] is not None and str(grid_val[r][c]).strip() != "")
                if non_empty_raw == 1 and non_empty_grid > 1:
                    banner_rows.add(r)

        # 3. Build occupancy grid and detect connected components
        occupancy = []
        for r in range(max_row):
            if r in banner_rows:
                occupancy.append([False for _ in range(max_column)])
            else:
                occupancy.append([(grid_val[r][c] is not None and str(grid_val[r][c]).strip() != "")
                                  for c in range(max_column)])
                      
        visited = set()
        components = []
        
        for r in range(max_row):
            for c in range(max_column):
                if occupancy[r][c] and (r, c) not in visited:
                    queue = [(r, c)]
                    visited.add((r, c))
                    comp = []
                    while queue:
                        curr = queue.pop(0)
                        comp.append(curr)
                        for dr in [-3, -2, -1, 0, 1, 2, 3]:
                            for dc in [-1, 0, 1]:
                                if dr == 0 and dc == 0:
                                    continue
                                nr, nc = curr[0] + dr, curr[1] + dc
                                if 0 <= nr < max_row and 0 <= nc < max_column:
                                    if occupancy[nr][nc] and (nr, nc) not in visited:
                                        visited.add((nr, nc))
                                        queue.append((nr, nc))
                    components.append(comp)

        raw_sheet_tables = []
        for comp in components:
            if len(comp) < 4:
                continue
                
            rows_in_comp = [cell[0] for cell in comp]
            cols_in_comp = [cell[1] for cell in comp]
            min_r, max_r = min(rows_in_comp), max(rows_in_comp)
            min_c, max_c = min(cols_in_comp), max(cols_in_comp)
            
            table_title_parts = []
            check_r = min_r - 1
            while check_r >= 0:
                if check_r in banner_rows:
                    for c_idx in range(max_column):
                        val = grid_val[check_r][c_idx]
                        if val is not None and str(val).strip() != "":
                            table_title_parts.insert(0, str(val).strip())
                            break
                    check_r -= 1
                else:
                    break
            table_title = " - ".join(table_title_parts) if table_title_parts else ""
            
            filtered_cols_indices = []
            for c_idx in range(min_c, max_c + 1):
                col_letter = openpyxl.utils.get_column_letter(c_idx + 1)
                if sheet_val.column_dimensions[col_letter].hidden:
                    continue
                filtered_cols_indices.append(c_idx)
                
            if not filtered_cols_indices:
                continue

            data_start_row = detect_data_start_row(sheet_val, min_r, max_r, filtered_cols_indices, grid_val)
            
            header = []
            for col_idx in filtered_cols_indices:
                col_header_parts = []
                for h_r in range(min_r, min_r + data_start_row):
                    val = grid_val[h_r][col_idx]
                    if val is not None and str(val).strip() != "" and str(val).strip().lower() != "nan":
                        col_header_parts.append(str(val).strip())
                if not col_header_parts:
                    header.append(f"col_{col_idx - min_c}")
                else:
                    header.append(" ".join(col_header_parts))
                
            final_columns = []
            seen = {}
            for h in header:
                h_clean = sanitize_name(h)
                if not h_clean:
                    h_clean = "column"
                if h_clean in seen:
                    seen[h_clean] += 1
                    final_columns.append(f"{h_clean}_{seen[h_clean]}")
                else:
                    seen[h_clean] = 1
                    final_columns.append(h_clean)
                    
            data_rows = []
            for r_idx in range(min_r + data_start_row, max_r + 1):
                if sheet_val.row_dimensions[r_idx + 1].hidden:
                    continue
                    
                row_data = []
                for c_idx in filtered_cols_indices:
                    val = grid_val[r_idx][c_idx]
                    form = grid_form[r_idx][c_idx]
                    if c_idx == min_c and c_idx > 0 and (val is None or str(val).strip() == ""):
                        first_col_val = grid_val[r_idx][0]
                        if first_col_val is not None and isinstance(first_col_val, str) and str(first_col_val).strip() != "":
                            val = first_col_val
                    if val is None and isinstance(form, str) and form.startswith("="):
                        val = form
                    row_data.append(val)
                data_rows.append(row_data)
                
            if not data_rows:
                continue
                
            df = pd.DataFrame(data_rows, columns=final_columns)
            if is_key_value_block(df):
                non_empty_cols = [col for col in df.columns if not df[col].isna().all()]
                kv_rows = []
                for _, row in df.iterrows():
                    k = row[non_empty_cols[0]]
                    v = row[non_empty_cols[1]]
                    if pd.isna(k) or str(k).strip() == "":
                        continue
                    k_str = str(k).strip()
                    if k_str.endswith(":"):
                        k_str = k_str[:-1].strip()
                    kv_rows.append([k_str, v])
                df = pd.DataFrame(kv_rows, columns=["key", "value"])
                final_columns = ["key", "value"]
            
            df = coerce_numeric_columns(df)
            
            raw_sheet_tables.append({
                "sheet_name": sheet_name,
                "table_title": table_title,
                "df": df,
                "final_columns": final_columns,
                "min_r": min_r,
                "max_r": max_r,
                "min_c": min_c,
                "max_c": max_c,
                "data_start_row": data_start_row,
                "filtered_cols_indices": filtered_cols_indices
            })

        # Phase 1: Fragment Merging within sheet
        merged_tables = []
        for t in raw_sheet_tables:
            merged = False
            for prev in merged_tables:
                if is_same_column_signature(prev["final_columns"], t["final_columns"]):
                    # Concatenate dataframes
                    try:
                        t_df = t["df"]
                        if list(t_df.columns) != list(prev["df"].columns):
                            t_df.columns = prev["df"].columns
                        prev["df"] = pd.concat([prev["df"], t_df], ignore_index=True)
                        prev["max_r"] = max(prev["max_r"], t["max_r"])
                        merged = True
                        logger.info(f"Merged table fragment ({len(t_df)} rows) into {prev['table_title'] or 'table'} in sheet '{sheet_name}'. Total rows: {len(prev['df'])}")
                        break
                    except Exception as merge_err:
                        logger.warning(f"Could not merge table fragments: {merge_err}")
            if not merged:
                merged_tables.append(t)

        # Phase 1 Part B: Absorb tiny fragments (<= 3 rows) into sheet_leftovers
        sheet_leftovers = []
        final_sheet_tables = []
        table_counter = 1
        for t in merged_tables:
            if len(t["df"]) <= 3 and len(merged_tables) > 1:
                # Treat as leftover footnote/note text chunk
                md_snippet = t["df"].to_markdown(index=False)
                sheet_leftovers.append(f"### Note / Context (Rows {t['min_r']+1}-{t['max_r']+1}):\n{md_snippet}")
            else:
                table_id = f"table_{table_counter}"
                table_counter += 1
                t["table_id"] = table_id
                t["db_table_name"] = sanitize_name(f"{filename}_{sheet_name}_{table_id}")
                final_sheet_tables.append(t)

        if sheet_leftovers:
            leftover_text = f"## File: {filename} | Sheet: {sheet_name} (Additional Context & Notes)\n\n" + "\n\n".join(sheet_leftovers)
            leftover_chunks.append({
                "text": leftover_text,
                "metadata": {
                    "source": file_path,
                    "source_name": filename,
                    "source_file": filename,
                    "type": "excel",
                    "sheet_name": sheet_name,
                    "table_id": f"{sheet_name}_leftovers",
                    "chunk_tier": "sheet_leftovers"
                }
            })

        all_sheet_tables.extend(final_sheet_tables)

    # Register all final merged tables with SQLite / Postgres
    for t in all_sheet_tables:
        sqlite_tables.append((t["db_table_name"], t["df"], t["table_title"]))

    # Phase 2: Deterministic Chunk Budget
    # Build Table Summary Chunks
    summary_chunks = []
    for t in all_sheet_tables:
        df = t["df"]
        db_table_name = t["db_table_name"]
        final_columns = t["final_columns"]
        table_title = t["table_title"]
        sheet_name = t["sheet_name"]
        
        preview_df = df.head(5).fillna("")
        preview_md = preview_df.to_markdown(index=False)
        title_str = f" (Title: {table_title})" if table_title else ""
        summary_text = (
            f"## File: {filename} | Sheet: {sheet_name}{title_str} (SQL Table: {db_table_name})\n"
            f"Total Rows: {len(df)}\n"
            f"Columns ({len(final_columns)}): {', '.join(final_columns)}\n\n"
            f"### Sample Rows Preview:\n{preview_md}"
        )
        dtypes_dict = {col: str(dtype) for col, dtype in df.dtypes.items()}
        summary_chunks.append({
            "text": summary_text,
            "metadata": {
                "source": file_path,
                "source_name": filename,
                "source_file": filename,
                "type": "excel",
                "sheet_name": sheet_name,
                "table_id": db_table_name,
                "row_range": [t["min_r"] + 1, t["max_r"] + 1],
                "columns": final_columns,
                "column_dtypes": dtypes_dict,
                "chunk_tier": "table_summary",
            }
        })

    chunks.extend(summary_chunks)
    chunks.extend(leftover_chunks)

    # Proportional window chunking
    total_summary_count = len(summary_chunks) + len(leftover_chunks)
    window_budget = max(0, 100 - total_summary_count)

    total_rows_all_tables = sum(len(t["df"]) for t in all_sheet_tables) or 1

    for t in all_sheet_tables:
        df = t["df"]
        nrows = len(df)
        if nrows <= 0 or window_budget <= 0:
            continue
            
        # Proportion of budget assigned to this table
        prop_windows = int((nrows / total_rows_all_tables) * window_budget)
        target_windows = min(max(prop_windows, 1 if nrows <= 50 else 0), 5)

        if target_windows > 0:
            step = max(nrows // target_windows, 15)
            window_indices = list(range(0, nrows, step))[:target_windows]
            for i in window_indices:
                window_df = df.iloc[i:i+15].fillna("")
                window_df = window_df.copy()
                for col in window_df.columns:
                    if window_df[col].dtype == object:
                        window_df[col] = window_df[col].astype(str).str.replace('\n', ' ', regex=False).str.replace('\r', '', regex=False)
                md_table = window_df.to_markdown(index=False)
                title_hdr = f" (Title: {t['table_title']})" if t['table_title'] else ""
                chunk_text = f"## Sheet: {t['sheet_name']}{title_hdr} (Table: {t['table_id']}, Rows {t['min_r'] + i + 1}-{t['min_r'] + min(i+15, nrows) + 1})\n\n{md_table}"
                dtypes_dict = {col: str(dtype) for col, dtype in df.dtypes.items()}
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source": file_path,
                        "source_name": filename,
                        "source_file": filename,
                        "type": "excel",
                        "sheet_name": t['sheet_name'],
                        "table_id": t['db_table_name'],
                        "row_range": [t['min_r'] + i + 1, t['min_r'] + min(i+15, nrows) + 1],
                        "columns": t['final_columns'],
                        "column_dtypes": dtypes_dict,
                        "chunk_tier": "markdown_window"
                    }
                })

    logger.info(f"Excel {filename}: {len(sqlite_tables)} merged table(s) → {len(chunks)} vector chunk(s) (all tables stored in SQL).")
    return chunks, sqlite_tables

def load_tables_to_sqlite(sqlite_tables: List[Tuple[str, pd.DataFrame, str]], company_id: str = None):
    """
    Persist parsed Tier-C tables for this tenant. Legacy name kept for the engine
    call sites, but it now dispatches (via table_store) to Postgres when
    VECTOR_BACKEND=pgvector, or SQLite for local dev. See backend/src/table_store.py.
    """
    from backend.src import table_store
    table_store.load_tables(sqlite_tables, company_id)

def delete_tables_from_sqlite(filename: str, company_id: str = None):
    """Delete all Tier-C tables for a filename (Postgres or SQLite; see table_store)."""
    from backend.src import table_store
    table_store.delete_tables_for_file(filename, company_id)

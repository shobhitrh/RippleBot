import os
import re
import json
import logging
import hashlib
from pathlib import Path
from datetime import datetime
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

class SmartExcelProcessor:
    """
    Dynamic, zero-hardcoding Excel & CSV processor for LLM RAG + SQL Pipelines.
    
    Dynamically analyzes sheet geometry, text density, merged cell ratios, and structural symmetry:
    1. Unmerges & forward-fills cells so header/category ranges never go blank.
    2. Auto-detects the true table header row (ignoring title/banner rows).
    3. Classifies sheets into 'TABULAR' (for SQL/Vector row chunks) vs 'DOCUMENT' (for Markdown/Section RAG chunks).
    4. For DOCUMENT sheets, converts to Markdown and chunks using specialized semantic Markdown block chunking.
    5. For TABULAR sheets, outputs clean Key-Value YAML Record Cards for Voyage-4-Large & Rerank-2.5.
    """
    
    def __init__(self, output_dir=None):
        self.output_dir = output_dir
        if output_dir:
            self.tabular_dir = os.path.join(output_dir, "tabular_sql")
            self.markdown_dir = os.path.join(output_dir, "markdown_docs")
            self.chunks_dir = os.path.join(output_dir, "rag_chunks")
            self.metadata_dir = os.path.join(output_dir, "metadata")
            for d in [self.tabular_dir, self.markdown_dir, self.chunks_dir, self.metadata_dir]:
                os.makedirs(d, exist_ok=True)

    @staticmethod
    def sanitize_name(name):
        clean = re.sub(r'[^\w]', '_', str(name)).strip()
        clean = re.sub(r'_+', '_', clean).lower().strip("_")
        if clean and clean[0].isdigit():
            clean = "t_" + clean
        return clean or "table"

    @staticmethod
    def unmerge_and_fill(ws):
        """
        openpyxl leaves merged cell ranges empty except the top-left one.
        Forward-fill the top-left value across the whole range before
        classification or extraction so merged headers/categories aren't lost.
        """
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            for merged_range in list(ws.merged_cells.ranges):
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_value = ws.cell(min_row, min_col).value
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception:
                    pass
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        ws.cell(row, col).value = top_left_value

    def detect_header_row(self, ws, min_r, max_r, min_c, max_c, scan_rows=5):
        """
        Auto-detect the true header row by scanning top non-empty rows for
        the highest string ratio and fill ratio, skipping title/spacer rows.
        """
        candidates = []
        scanned = 0
        r = min_r
        while r <= max_r and scanned < scan_rows:
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            non_empty = [v for v in row_vals if v is not None and str(v).strip() != '']
            if non_empty:
                str_ratio = sum(1 for v in non_empty if isinstance(v, str)) / len(non_empty)
                fill_ratio = len(non_empty) / max(max_c - min_c + 1, 1)
                candidates.append((r, str_ratio, fill_ratio))
                scanned += 1
            r += 1

        if not candidates:
            return min_r

        # Prefer high string ratio and high fill ratio
        best = sorted(candidates, key=lambda x: (-x[1], -x[2], x[0]))[0]
        return best[0]

    def analyze_sheet_type(self, ws, min_r, max_r, min_c, max_c):
        """
        Dynamic Heuristic Classifier (Zero Hardcoding):
        Determines if a sheet is 'DOCUMENT' or 'TABULAR' based on structural metrics.
        """
        total_rows = max_r - min_r + 1
        total_cols = max_c - min_c + 1
        
        if total_rows <= 1 or total_cols <= 1:
            return "DOCUMENT"
            
        merged_count = len(getattr(ws, 'merged_cells', {}).ranges) if hasattr(ws, 'merged_cells') and getattr(ws, 'merged_cells') else 0
        merged_ratio = merged_count / max(total_rows, 1)
        
        # Analyze row fill profiles
        row_lengths = []
        text_heavy_rows = 0
        for r in range(min_r, max_r + 1):
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            non_empty = [v for v in row_vals if v is not None and str(v).strip() != '']
            row_lengths.append(len(non_empty))
            
            # Check if row is a single long narrative text string
            if len(non_empty) == 1 and isinstance(non_empty[0], str) and len(non_empty[0]) > 40:
                text_heavy_rows += 1
                
        avg_len = sum(row_lengths) / len(row_lengths) if row_lengths else 0
        variance = sum((l - avg_len) ** 2 for l in row_lengths) / len(row_lengths) if row_lengths else 0
        text_heavy_ratio = text_heavy_rows / max(total_rows, 1)
        
        # TABULAR condition: High row uniformity, low text-heavy narrative rows, low merged ratio
        if variance < 4.0 and text_heavy_ratio < 0.15 and merged_ratio < 0.2 and total_rows >= 5:
            return "TABULAR"
        else:
            return "DOCUMENT"

    @staticmethod
    def split_markdown_blocks(md: str) -> list:
        """Split markdown text into semantic blocks (headers, code blocks, tables, lists)."""
        blocks = []
        buf = []
        in_code = False

        lines = md.splitlines()
        for line in lines:
            if line.strip().startswith("```"):
                in_code = not in_code
                buf.append(line)
                continue

            if in_code:
                buf.append(line)
                continue

            if re.match(r"^#{1,6}\s+", line):
                if buf:
                    blocks.append("\n".join(buf).strip())
                    buf = []
                buf.append(line)
                continue

            if "|" in line and re.match(r"^\s*\|.*\|\s*$", line):
                buf.append(line)
                continue

            if re.match(r"^\s*[-*+]\s+", line):
                buf.append(line)
                continue

            if line.strip() == "":
                if buf:
                    blocks.append("\n".join(buf).strip())
                    buf = []
                continue

            buf.append(line)

        if buf:
            blocks.append("\n".join(buf).strip())

        return [b for b in blocks if b.strip()]

    def chunk_markdown_blocks(self, blocks: list, max_tokens: int = 800, overlap_tokens: int = 120) -> list:
        """Chunk markdown blocks with token budget control and overlap."""
        try:
            import tiktoken
            enc = tiktoken.get_encoding("cl100k_base")
            def count_tok(t): return len(enc.encode(t))
            def decode_tok(toks): return enc.decode(toks)
            def encode_tok(t): return enc.encode(t)
        except Exception:
            def count_tok(t): return max(1, len(t) // 4)
            def decode_tok(toks): return "".join(toks)
            def encode_tok(t): return list(t)

        chunks = []
        current = []
        current_tokens = 0

        def flush_with_overlap():
            nonlocal current, current_tokens
            if not current:
                return
            chunk_text = "\n\n".join(current).strip()
            chunks.append(chunk_text)

            if overlap_tokens > 0:
                tokens = encode_tok(chunk_text)
                overlap = tokens[-overlap_tokens:] if len(tokens) > overlap_tokens else tokens
                overlap_text = decode_tok(overlap)
                current = [overlap_text]
                current_tokens = count_tok(overlap_text)
            else:
                current = []
                current_tokens = 0

        for block in blocks:
            block_tokens = count_tok(block)

            if block_tokens > max_tokens:
                if current:
                    flush_with_overlap()

                tokens = encode_tok(block)
                start = 0
                while start < len(tokens):
                    end = min(start + max_tokens, len(tokens))
                    chunk = decode_tok(tokens[start:end])
                    chunks.append(chunk)
                    start = end - overlap_tokens if overlap_tokens > 0 and end - overlap_tokens > start else end
                current = []
                current_tokens = 0
                continue

            if current_tokens + block_tokens <= max_tokens:
                current.append(block)
                current_tokens += block_tokens
            else:
                flush_with_overlap()
                current.append(block)
                current_tokens += block_tokens

        if current:
            flush_with_overlap()

        return chunks

    def process_document_sheet(self, ws, file_name, sheet_name, min_r, max_r, min_c, max_c, file_path=None):
        """
        Parses mixed/document sheets into clean Markdown & chunks using specialized Markdown block chunking.
        """
        md_lines = [f"# {file_name} - {sheet_name}\n"]
        current_section = "General"

        for r in range(min_r, max_r + 1):
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            non_empty = [str(v).strip() for v in row_vals if v is not None and str(v).strip() != '']

            if not non_empty:
                continue

            # Detect section headers (single string, short, not ending in period)
            if len(non_empty) == 1:
                text = non_empty[0]
                if len(text) < 70 and not text.endswith('.'):
                    current_section = text
                    md_lines.append(f"\n## {text}\n")
                else:
                    md_lines.append(f"- {text}")
            else:
                line = "| " + " | ".join([t.replace('|', '&#124;').replace('\n', ' ') for t in non_empty]) + " |"
                md_lines.append(line)

        md_content = "\n".join(md_lines) + "\n"

        # Apply specialized Markdown chunker
        blocks = self.split_markdown_blocks(md_content)
        chunk_texts = self.chunk_markdown_blocks(blocks, max_tokens=800, overlap_tokens=120)

        chunks = []
        for i, text in enumerate(chunk_texts):
            header_match = re.search(r'^#{1,6}\s+(.+)$', text, re.MULTILINE)
            section = header_match.group(1) if header_match else current_section

            chunks.append({
                "text": f"[{file_name} > {sheet_name}]\n{text}",
                "metadata": {
                    "source": file_path or file_name,
                    "source_name": file_name,
                    "file_name": file_name,
                    "sheet": sheet_name,
                    "sheet_name": sheet_name,
                    "section": section,
                    "type": "document_section",
                    "chunk_index": i,
                    "indexed_at": datetime.now().isoformat()
                }
            })

        return md_content, chunks

    def process_tabular_sheet(self, ws, file_name, sheet_name, header_r, max_r, min_c, max_c, file_path=None):
        """
        Parses uniform data grids into SQL DataFrames & YAML Record Chunks for RAG.
        """
        if hasattr(ws, 'cell'):
            raw_headers = [ws.cell(header_r, c).value for c in range(min_c, max_c + 1)]
        else:
            raw_headers = ws[header_r - 1]
            
        clean_headers = []
        seen = {}
        for idx, h in enumerate(raw_headers):
            h_str = str(h).strip() if h is not None and str(h).strip() != '' else f"col_{idx+1}"
            h_clean = self.sanitize_name(h_str)
            if h_clean in seen:
                seen[h_clean] += 1
                h_clean = f"{h_clean}_{seen[h_clean]}"
            else:
                seen[h_clean] = 1
            clean_headers.append(h_clean)
            
        data_rows = []
        chunks = []
        
        data_start_r = header_r + 1
        for r_idx, r in enumerate(range(data_start_r, max_r + 1), start=1):
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            if not any(v is not None and str(v).strip() != '' for v in row_vals):
                continue
                
            data_rows.append(row_vals)
            
            # Format row as self-contained YAML record for Voyage-4 Embedding & Neon Vector RAG
            yaml_lines = [f"# Source: {file_name} | Sheet: {sheet_name} | Row: {r_idx}"]
            row_dict = {}
            for h, v in zip(clean_headers, row_vals):
                if v is not None and str(v).strip() != '':
                    v_clean = str(v).replace('\n', ' ').strip()
                    yaml_lines.append(f"{h}: {v_clean}")
                    row_dict[h] = v_clean
                    
            chunk_text = "\n".join(yaml_lines)
            chunks.append({
                "text": chunk_text,
                "metadata": {
                    "source": file_path or file_name,
                    "source_name": file_name,
                    "file_name": file_name,
                    "sheet": sheet_name,
                    "sheet_name": sheet_name,
                    "row_index": r_idx,
                    "type": "record_card",
                    "chunk_index": r_idx - 1,
                    "indexed_at": datetime.now().isoformat(),
                    **row_dict
                }
            })
            
        df = pd.DataFrame(data_rows, columns=clean_headers)
        return df, chunks

    def process_file(self, file_path):
        file_name = os.path.basename(file_path)
        file_stem = self.sanitize_name(os.path.splitext(file_name)[0])
        ext = os.path.splitext(file_name)[1].lower()
        
        logger.info(f"Smart Processing File: {file_name}")
        
        all_chunks = []
        sqlite_tables = []
        
        if ext in ('.csv', '.tsv'):
            sep = '\t' if ext == '.tsv' else ','
            try:
                df_raw = pd.read_csv(file_path, sep=sep, header=None)
                df_raw = df_raw.dropna(how='all')
                if not df_raw.empty:
                    rows = df_raw.values.tolist()
                    min_r, max_r = 1, len(rows)
                    min_c, max_c = 1, max(len(r) for r in rows)
                    sheet_name = "Sheet1"
                    header_r = self.detect_header_row(rows, min_r, max_r, min_c, max_c)
                    sheet_type = self.analyze_sheet_type(rows, min_r, max_r, min_c, max_c)
                    if sheet_type == "TABULAR":
                        df, sheet_chunks = self.process_tabular_sheet(rows, file_name, sheet_name, header_r, max_r, min_c, max_c, file_path=file_path)
                        table_name = f"{file_stem}_{self.sanitize_name(sheet_name)}_table_1"
                        sqlite_tables.append((table_name, df, f"{file_stem} - {sheet_name}"))
                        all_chunks.extend(sheet_chunks)
                    else:
                        md_content, sheet_chunks = self.process_document_sheet(rows, file_name, sheet_name, min_r, max_r, min_c, max_c, file_path=file_path)
                        all_chunks.extend(sheet_chunks)
            except Exception as e:
                logger.error(f"Smart Excel Processor failed for CSV {file_name}: {e}")
            return all_chunks, sqlite_tables

        # Excel processing (.xlsx, .xls)
        try:
            wb_val = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"openpyxl failed to load {file_path}: {e}")
            return all_chunks, sqlite_tables
            
        for sname in wb_val.sheetnames:
            ws_v = wb_val[sname]
            sanitized_sheet = self.sanitize_name(sname)
            
            # Resolve merged cells before doing layout bounds & extraction
            self.unmerge_and_fill(ws_v)
            
            all_rows = list(ws_v.iter_rows(values_only=True))
            min_r, max_r = 1000000, 0
            min_c, max_c = 1000000, 0
            for r_idx, row in enumerate(all_rows, start=1):
                for c_idx, val in enumerate(row, start=1):
                    if val is not None and str(val).strip() != '':
                        if r_idx < min_r: min_r = r_idx
                        if r_idx > max_r: max_r = r_idx
                        if c_idx < min_c: min_c = c_idx
                        if c_idx > max_c: max_c = c_idx
                        
            if max_r == 0:
                logger.info(f"Sheet '{sname}': Skipping Empty Sheet")
                continue
                
            header_r = self.detect_header_row(ws_v, min_r, max_r, min_c, max_c)
            sheet_type = self.analyze_sheet_type(ws_v, min_r, max_r, min_c, max_c)
            logger.info(f"Sheet '{sname}' -> Classified as: [{sheet_type}] (Header Row: {header_r}, Rows: {max_r-min_r+1}, Cols: {max_c-min_c+1})")
            
            if sheet_type == "TABULAR":
                df, sheet_chunks = self.process_tabular_sheet(ws_v, file_name, sname, header_r, max_r, min_c, max_c, file_path=file_path)
                table_name = f"{file_stem}_{sanitized_sheet}_table_1"
                sqlite_tables.append((table_name, df, f"{file_stem} - {sname}"))
                all_chunks.extend(sheet_chunks)
            else:
                md_content, sheet_chunks = self.process_document_sheet(ws_v, file_name, sname, min_r, max_r, min_c, max_c, file_path=file_path)
                all_chunks.extend(sheet_chunks)

        return all_chunks, sqlite_tables
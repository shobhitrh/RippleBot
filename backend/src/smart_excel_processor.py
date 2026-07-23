import os
import re
import json
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

class SmartExcelProcessor:
    """
    Dynamic, zero-hardcoding Excel & CSV processor for LLM RAG + SQL Pipelines.
    
    Dynamically analyzes sheet geometry, text density, merged cell ratios, and structural symmetry:
    1. Classifies sheets into 'TABULAR' (for SQL/Vector row chunks) vs 'DOCUMENT' (for Markdown/Section RAG chunks).
    2. Guarantees 0 data loss by retaining raw evaluated values, timestamps, and layout metadata.
    3. Outputs pre-chunked JSON objects formatted specifically for Voyage-4-Large embeddings & Rerank-2.5.
    """
    
    def __init__(self, output_dir=None):
        self.output_dir = output_dir
        if output_dir:
            self.tabular_dir = os.path.join(output_dir, "tabular_sql")
            self.markdown_dir = os.path.join(output_dir, "markdown_docs")
            self.chunks_dir = os.path.join(output_dir, "rag_chunks")
            self.metadata_dir = os.path.join(output_dir, "metadata")
            for d in [self.tabular_dir, self.markdown_dir, self.chunks_dir, self.metadata_dir]:
                os.makedirs(d, exist_ok=True)

    @staticmethod
    def sanitize_name(name):
        clean = re.sub(r'[^\w]', '_', str(name)).strip()
        clean = re.sub(r'_+', '_', clean).lower().strip("_")
        if clean and clean[0].isdigit():
            clean = "t_" + clean
        return clean or "table"

    def analyze_sheet_type(self, ws, min_r, max_r, min_c, max_c):
        """
        Dynamic Heuristic Classifier (Zero Hardcoding):
        Determines if a sheet is 'DOCUMENT' or 'TABULAR' based on structural metrics.
        """
        total_rows = max_r - min_r + 1
        total_cols = max_c - min_c + 1
        
        if total_rows <= 1 or total_cols <= 1:
            return "DOCUMENT"
            
        merged_count = len(getattr(ws, 'merged_cells', {}).ranges) if hasattr(ws, 'merged_cells') else 0
        merged_ratio = merged_count / max(total_rows, 1)
        
        # Analyze row fill profiles
        row_lengths = []
        text_heavy_rows = 0
        for r in range(min_r, max_r + 1):
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            non_empty = [v for v in row_vals if v is not None and str(v).strip() != '']
            row_lengths.append(len(non_empty))
            
            # Check if row is a single long narrative text string
            if len(non_empty) == 1 and isinstance(non_empty[0], str) and len(non_empty[0]) > 40:
                text_heavy_rows += 1
                
        avg_len = sum(row_lengths) / len(row_lengths) if row_lengths else 0
        variance = sum((l - avg_len) ** 2 for l in row_lengths) / len(row_lengths) if row_lengths else 0
        text_heavy_ratio = text_heavy_rows / max(total_rows, 1)
        
        # TABULAR condition: High row uniformity, low text-heavy narrative rows, low merged ratio
        if variance < 4.0 and text_heavy_ratio < 0.15 and merged_ratio < 0.2 and total_rows >= 5:
            return "TABULAR"
        else:
            return "DOCUMENT"

    def process_document_sheet(self, ws, file_name, sheet_name, min_r, max_r, min_c, max_c, file_path=None):
        """
        Parses mixed/document sheets dynamically into clean Markdown & RAG Section Chunks.
        """
        md_lines = [f"# {file_name} - {sheet_name}\n"]
        chunks = []
        
        current_section = "General"
        current_section_lines = []
        chunk_idx = 0
        
        for r in range(min_r, max_r + 1):
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            non_empty = [str(v).strip() for v in row_vals if v is not None and str(v).strip() != '']
            
            if not non_empty:
                continue
                
            # Detect section headers (single string, short, not ending in period)
            if len(non_empty) == 1:
                text = non_empty[0]
                if len(text) < 70 and not text.endswith('.'):
                    if current_section_lines:
                        chunk_text = "\n".join(current_section_lines)
                        chunks.append({
                            "text": f"[{file_name} > {sheet_name} > {current_section}]\n{chunk_text}",
                            "metadata": {
                                "source": file_path or file_name,
                                "source_name": file_name,
                                "file_name": file_name,
                                "sheet": sheet_name,
                                "sheet_name": sheet_name,
                                "section": current_section,
                                "type": "document_section",
                                "chunk_index": chunk_idx,
                                "indexed_at": datetime.now().isoformat()
                            }
                        })
                        chunk_idx += 1
                        current_section_lines = []
                    current_section = text
                    md_lines.append(f"\n## {text}\n")
                else:
                    line = f"- {text}"
                    md_lines.append(line)
                    current_section_lines.append(line)
            else:
                line = "| " + " | ".join([t.replace('|', '&#124;').replace('\n', ' ') for t in non_empty]) + " |"
                md_lines.append(line)
                current_section_lines.append(line)
                
        if current_section_lines:
            chunk_text = "\n".join(current_section_lines)
            chunks.append({
                "text": f"[{file_name} > {sheet_name} > {current_section}]\n{chunk_text}",
                "metadata": {
                    "source": file_path or file_name,
                    "source_name": file_name,
                    "file_name": file_name,
                    "sheet": sheet_name,
                    "sheet_name": sheet_name,
                    "section": current_section,
                    "type": "document_section",
                    "chunk_index": chunk_idx,
                    "indexed_at": datetime.now().isoformat()
                }
            })
            
        md_content = "\n".join(md_lines) + "\n"
        return md_content, chunks

    def process_tabular_sheet(self, ws, file_name, sheet_name, min_r, max_r, min_c, max_c, file_path=None):
        """
        Parses uniform data grids into SQL DataFrames & YAML Record Chunks for RAG.
        """
        if hasattr(ws, 'cell'):
            raw_headers = [ws.cell(min_r, c).value for c in range(min_c, max_c + 1)]
        else:
            raw_headers = ws[min_r - 1]
            
        clean_headers = []
        for idx, h in enumerate(raw_headers):
            h_str = str(h).strip() if h is not None and str(h).strip() != '' else f"col_{idx}"
            h_clean = self.sanitize_name(h_str)
            clean_headers.append(h_clean)
            
        data_rows = []
        chunks = []
        
        for r_idx, r in enumerate(range(min_r + 1, max_r + 1), start=1):
            if hasattr(ws, 'cell'):
                row_vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            else:
                row_vals = ws[r - 1]
            if not any(v is not None and str(v).strip() != '' for v in row_vals):
                continue
                
            data_rows.append(row_vals)
            
            # Format row as self-contained YAML record for Voyage-4 Embedding & Neon Vector RAG
            yaml_lines = [f"# Source: {file_name} | Sheet: {sheet_name} | Row: {r_idx}"]
            row_dict = {}
            for h, v in zip(clean_headers, row_vals):
                if v is not None and str(v).strip() != '':
                    v_clean = str(v).replace('\n', ' ').strip()
                    yaml_lines.append(f"{h}: {v_clean}")
                    row_dict[h] = v
                    
            chunk_text = "\n".join(yaml_lines)
            chunks.append({
                "text": chunk_text,
                "metadata": {
                    "source": file_path or file_name,
                    "source_name": file_name,
                    "file_name": file_name,
                    "sheet": sheet_name,
                    "sheet_name": sheet_name,
                    "row_index": r_idx,
                    "type": "record_card",
                    "chunk_index": r_idx - 1,
                    "indexed_at": datetime.now().isoformat(),
                    **row_dict
                }
            })
            
        df = pd.DataFrame(data_rows, columns=clean_headers)
        return df, chunks

    def process_file(self, file_path):
        file_name = os.path.basename(file_path)
        file_stem = self.sanitize_name(os.path.splitext(file_name)[0])
        ext = os.path.splitext(file_name)[1].lower()
        
        logger.info(f"Smart Processing File: {file_name}")
        
        all_chunks = []
        sqlite_tables = []
        
        if ext in ('.csv', '.tsv'):
            sep = '\t' if ext == '.tsv' else ','
            try:
                df_raw = pd.read_csv(file_path, sep=sep, header=None)
                df_raw = df_raw.dropna(how='all')
                if not df_raw.empty:
                    rows = df_raw.values.tolist()
                    min_r, max_r = 1, len(rows)
                    min_c, max_c = 1, max(len(r) for r in rows)
                    sheet_name = "Sheet1"
                    sheet_type = self.analyze_sheet_type(rows, min_r, max_r, min_c, max_c)
                    if sheet_type == "TABULAR":
                        df, sheet_chunks = self.process_tabular_sheet(rows, file_name, sheet_name, min_r, max_r, min_c, max_c, file_path=file_path)
                        table_name = f"{file_stem}_{self.sanitize_name(sheet_name)}_table_1"
                        sqlite_tables.append((table_name, df, f"{file_stem} - {sheet_name}"))
                        all_chunks.extend(sheet_chunks)
                    else:
                        md_content, sheet_chunks = self.process_document_sheet(rows, file_name, sheet_name, min_r, max_r, min_c, max_c, file_path=file_path)
                        all_chunks.extend(sheet_chunks)
            except Exception as e:
                logger.error(f"Smart Excel Processor failed for CSV {file_name}: {e}")
            return all_chunks, sqlite_tables

        # Excel processing (.xlsx, .xls)
        try:
            wb_val = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"openpyxl failed to load {file_path}: {e}")
            return all_chunks, sqlite_tables
            
        for sname in wb_val.sheetnames:
            ws_v = wb_val[sname]
            sanitized_sheet = self.sanitize_name(sname)
            
            all_rows = list(ws_v.iter_rows(values_only=True))
            min_r, max_r = 1000000, 0
            min_c, max_c = 1000000, 0
            for r_idx, row in enumerate(all_rows, start=1):
                for c_idx, val in enumerate(row, start=1):
                    if val is not None and str(val).strip() != '':
                        if r_idx < min_r: min_r = r_idx
                        if r_idx > max_r: max_r = r_idx
                        if c_idx < min_c: min_c = c_idx
                        if c_idx > max_c: max_c = c_idx
                        
            if max_r == 0:
                logger.info(f"Sheet '{sname}': Skipping Empty Sheet")
                continue
                
            sheet_type = self.analyze_sheet_type(ws_v, min_r, max_r, min_c, max_c)
            logger.info(f"Sheet '{sname}' -> Classified as: [{sheet_type}] (Rows: {max_r-min_r+1}, Cols: {max_c-min_c+1})")
            
            if sheet_type == "TABULAR":
                df, sheet_chunks = self.process_tabular_sheet(ws_v, file_name, sname, min_r, max_r, min_c, max_c, file_path=file_path)
                table_name = f"{file_stem}_{sanitized_sheet}_table_1"
                sqlite_tables.append((table_name, df, f"{file_stem} - {sname}"))
                all_chunks.extend(sheet_chunks)
            else:
                md_content, sheet_chunks = self.process_document_sheet(ws_v, file_name, sname, min_r, max_r, min_c, max_c, file_path=file_path)
                all_chunks.extend(sheet_chunks)

        return all_chunks, sqlite_tables

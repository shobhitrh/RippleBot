import os
import re
import gzip
import json
import logging
from pathlib import Path
from datetime import datetime
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

try:
    import tiktoken
    _ENC = tiktoken.get_encoding("cl100k_base")
    def count_tokens(text):
        return len(_ENC.encode(text))
except Exception:
    def count_tokens(text):
        return max(1, len(text) // 4)


class UnifiedExcelRAGPipeline:
    """
    All-in-One Enterprise Pipeline for Excel & CSV RAG:
    ---------------------------------------------------
    Stage 1: Unified Markdown Generator (0 Data Loss, Clean Unmerge, Formula Capture)
    Stage 2: Gzip Archive Compressor (.md.gz, 80-90% disk reduction)
    Stage 3: Header-Injected RAG Chunker (TOC + Breadcrumbs + Table Headers for Voyage Reranker)
    """

    def __init__(
        self,
        output_dir="unified_rag_pipeline_output",
        target_chunk_tokens=400,
        max_chunk_tokens=550
    ):
        self.output_dir = output_dir
        self.md_dir = os.path.join(output_dir, "unified_md")
        self.gz_dir = os.path.join(output_dir, "compressed_archives_gz")
        self.chunks_dir = os.path.join(output_dir, "rag_chunks_json")
        self.target_tokens = target_chunk_tokens
        self.max_tokens = max_chunk_tokens

        for d in [self.md_dir, self.gz_dir, self.chunks_dir]:
            os.makedirs(d, exist_ok=True)

    @staticmethod
    def sanitize_name(name):
        clean = re.sub(r'[^\w]', '_', str(name)).strip()
        clean = re.sub(r'_+', '_', clean).lower().strip("_")
        if clean and clean[0].isdigit():
            clean = "t_" + clean
        return clean or "table"

    @staticmethod
    def unmerge_cells_cleanly(ws_val, ws_form=None):
        """Unmerge cells: Top-left cell keeps value/formula, continuation cells set to None."""
        if hasattr(ws_val, 'merged_cells') and ws_val.merged_cells:
            for merged_range in list(ws_val.merged_cells.ranges):
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_v = ws_val.cell(min_row, min_col).value
                top_f = ws_form.cell(min_row, min_col).value if ws_form else None
                
                try: ws_val.unmerge_cells(str(merged_range))
                except Exception: pass
                if ws_form:
                    try: ws_form.unmerge_cells(str(merged_range))
                    except Exception: pass
                    
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if r == min_row and c == min_col:
                            ws_val.cell(r, c).value = top_v
                            if ws_form: ws_form.cell(r, c).value = top_f
                        else:
                            ws_val.cell(r, c).value = None
                            if ws_form: ws_form.cell(r, c).value = None

    # STAGE 1: Convert Excel Workbook / CSV to Unified Markdown + SQL DataFrames
    def generate_unified_markdown(self, file_path):
        file_name = os.path.basename(file_path)
        file_stem = self.sanitize_name(os.path.splitext(file_name)[0])
        ext = os.path.splitext(file_name)[1].lower()

        sqlite_tables = []
        md_lines = [f"# UNIFIED WORKBOOK: {file_name}\n"]

        if ext in ('.csv', '.tsv'):
            sep = '\t' if ext == '.tsv' else ','
            try:
                df = pd.read_csv(file_path, sep=sep)
                df = df.dropna(how='all')
                sheet_name = "Sheet1"
                table_name = f"{file_stem}_{self.sanitize_name(sheet_name)}_table_1"
                sqlite_tables.append((table_name, df, f"{file_stem} - {sheet_name}"))

                md_lines.append(f"\n---\n## Sheet: {sheet_name}\n")
                if not df.empty:
                    headers = [str(c).replace('\n', ' ').replace('|', '&#124;').strip() for c in df.columns]
                    md_lines.append("| " + " | ".join(headers) + " |")
                    md_lines.append("| " + " | ".join([":---"] * len(headers)) + " |")
                    for _, row in df.iterrows():
                        cells = [str(val).replace('\n', ' ').replace('|', '&#124;').strip() if pd.notna(val) else "-" for val in row]
                        md_lines.append("| " + " | ".join(cells) + " |")
            except Exception as e:
                logger.error(f"Failed to convert CSV {file_name} to Markdown: {e}")
        else:
            try:
                wb_v = openpyxl.load_workbook(file_path, data_only=True)
                wb_f = openpyxl.load_workbook(file_path, data_only=False)
            except Exception as e:
                logger.error(f"openpyxl failed to load {file_path}: {e}")
                return None, sqlite_tables

            for sname in wb_v.sheetnames:
                ws_v, ws_f = wb_v[sname], wb_f[sname]
                self.unmerge_cells_cleanly(ws_v, ws_f)
                sanitized_sheet = self.sanitize_name(sname)

                rows_v = list(ws_v.iter_rows(values_only=True))
                min_r, max_r, min_c, max_c = 1000000, 0, 1000000, 0
                for r_idx, row in enumerate(rows_v, start=1):
                    for c_idx, val in enumerate(row, start=1):
                        if val is not None and str(val).strip() != '':
                            min_r = min(min_r, r_idx); max_r = max(max_r, r_idx)
                            min_c = min(min_c, c_idx); max_c = max(max_c, c_idx)

                if max_r == 0:
                    md_lines.append(f"\n---\n## Sheet: {sname}\n*(Empty Sheet)*\n")
                    continue

                md_lines.append(f"\n---\n## Sheet: {sname}\n")
                in_table = False
                table_grid = []
                grid_headers = []

                for r in range(min_r, max_r + 1):
                    rv = [ws_v.cell(r, c).value for c in range(min_c, max_c + 1)]
                    rf = [ws_f.cell(r, c).value for c in range(min_c, max_c + 1)]
                    non_empty = [v for v in rv if v is not None and str(v).strip() != '']

                    if not non_empty:
                        in_table = False
                        continue

                    if len(non_empty) == 1 and isinstance(non_empty[0], str):
                        text = non_empty[0].strip()
                        in_table = False
                        is_bullet = (
                            text.startswith('-') or
                            text.lower().startswith('terms') or
                            re.search(r'\b(rs|inr|\$|\%|per|min|mins|multiple|multiples|upfront|net|taxes)\b', text.lower()) or
                            text.endswith('.')
                        )
                        if not is_bullet and len(text) < 60:
                            md_lines.append(f"\n### {text}\n")
                        else:
                            clean_bullet = text if text.startswith('-') else f"- {text}"
                            md_lines.append(clean_bullet)
                    else:
                        cells = []
                        raw_cells = []
                        for cv, cf in zip(rv, rf):
                            if cv is None:
                                cells.append("-")
                                raw_cells.append("")
                            else:
                                v_str = str(cv).replace('\n', ' ').replace('|', '&#124;').strip()
                                raw_cells.append(v_str)
                                if isinstance(cf, str) and cf.startswith('='):
                                    v_str = f"{v_str} (formula: {cf})"
                                cells.append(v_str if v_str != '' else "-")

                        if not in_table:
                            md_lines.append("\n| " + " | ".join(cells) + " |")
                            md_lines.append("| " + " | ".join([":---"] * len(cells)) + " |")
                            in_table = True
                            grid_headers = [c if c != '-' else f"col_{idx+1}" for idx, c in enumerate(raw_cells)]
                            table_grid = []
                        else:
                            md_lines.append("| " + " | ".join(cells) + " |")
                            table_grid.append(raw_cells)

                # Extract Tier-C SQL DataFrame for this sheet if tabular grid exists
                if grid_headers and table_grid:
                    try:
                        clean_cols = []
                        seen = {}
                        for idx, h in enumerate(grid_headers):
                            hc = self.sanitize_name(h) or f"col_{idx+1}"
                            if hc in seen:
                                seen[hc] += 1
                                hc = f"{hc}_{seen[hc]}"
                            else:
                                seen[hc] = 1
                            clean_cols.append(hc)
                        df_sheet = pd.DataFrame(table_grid, columns=clean_cols)
                        table_name = f"{file_stem}_{sanitized_sheet}_table_1"
                        sqlite_tables.append((table_name, df_sheet, f"{file_stem} - {sname}"))
                    except Exception as ex:
                        logger.warning(f"Could not extract SQL dataframe for sheet {sname}: {ex}")

        md_path = os.path.join(self.md_dir, f"{file_stem}_UNIFIED.md")
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(md_lines) + "\n")

        return md_path, sqlite_tables

    # STAGE 2: Compress Markdown to .md.gz
    def compress_to_gz(self, md_path):
        gz_path = os.path.join(self.gz_dir, os.path.basename(md_path) + ".gz")
        with open(md_path, 'rb') as f_in, gzip.open(gz_path, 'wb') as f_out:
            f_out.writelines(f_in)
        return gz_path

    # STAGE 3: Stream .md.gz archive and Chunk for RAG / Voyage Reranker
    def chunk_compressed_archive(self, gz_path, file_name, file_path=None):
        with gzip.open(gz_path, "rt", encoding="utf-8") as f:
            md_content = f.read()

        file_stem = self.sanitize_name(os.path.splitext(file_name)[0])
        chunks = []

        # 1. File Table of Contents (TOC) Chunk for reranker routing
        sheets = re.findall(r'^##\s+Sheet:\s+(.+)$', md_content, re.MULTILINE)
        toc_text = f"# File Summary & Table of Contents: {file_name}\nAvailable Sheets ({len(sheets)}):\n"
        for s in sheets:
            toc_text += f"- Sheet: {s}\n"

        chunks.append({
            "text": toc_text,
            "metadata": {
                "source": file_path or file_name,
                "source_name": file_name,
                "file_name": file_name,
                "type": "file_toc",
                "chunk_index": 0,
                "indexed_at": datetime.now().isoformat()
            }
        })

        # 2. Header-Injected Markdown Chunking
        sheets_raw = re.split(r'\n---\n##\s+Sheet:\s+', md_content)
        chunk_idx = 1

        for s_block in sheets_raw[1:]:
            s_lines = s_block.splitlines()
            sname = s_lines[0].strip() if s_lines else "General"
            sections_raw = re.split(r'\n###\s+', "\n".join(s_lines[1:]).strip())

            for sec_block in sections_raw:
                if not sec_block.strip(): continue

                sec_lines = sec_block.splitlines()
                sec_title = sec_lines[0].strip() if len(sec_lines) > 1 and not sec_lines[0].startswith('|') and not sec_lines[0].startswith('-') else "General"

                table_header = None
                data_lines = []

                for line in sec_lines:
                    if line.startswith('|') and not line.startswith('| :---'):
                        if table_header is None:
                            table_header = line
                        else:
                            data_lines.append(line)
                    elif line.strip() != '' and not line.startswith('| :---'):
                        data_lines.append(line)

                breadcrumb = f"# Source: {file_name} | Sheet: {sname} | Section: {sec_title}\n"
                batch = [breadcrumb]
                if table_header:
                    batch.append(table_header)
                    batch.append("| " + " | ".join([":---"] * table_header.count('|')) + " |")

                curr_toks = count_tokens("\n".join(batch))

                for line in data_lines:
                    line_toks = count_tokens(line)
                    if curr_toks + line_toks > self.max_tokens:
                        chunk_txt = "\n".join(batch)
                        chunks.append({
                            "text": chunk_txt,
                            "metadata": {
                                "source": file_path or file_name,
                                "source_name": file_name,
                                "file_name": file_name,
                                "sheet": sname,
                                "section": sec_title,
                                "type": "section_data",
                                "chunk_index": chunk_idx,
                                "indexed_at": datetime.now().isoformat()
                            }
                        })
                        chunk_idx += 1
                        batch = [breadcrumb]
                        if table_header: batch.append(table_header)
                        batch.append(line)
                        curr_toks = count_tokens("\n".join(batch))
                    else:
                        batch.append(line)
                        curr_toks += line_toks

                if len(batch) > (2 if table_header else 1):
                    chunk_txt = "\n".join(batch)
                    chunks.append({
                        "text": chunk_txt,
                        "metadata": {
                            "source": file_path or file_name,
                            "source_name": file_name,
                            "file_name": file_name,
                            "sheet": sname,
                            "section": sec_title,
                            "type": "section_data",
                            "chunk_index": chunk_idx,
                            "indexed_at": datetime.now().isoformat()
                        }
                    })
                    chunk_idx += 1

        chunks_path = os.path.join(self.chunks_dir, f"{file_stem}_RAG_CHUNKS.json")
        with open(chunks_path, 'w', encoding='utf-8') as f:
            json.dump(chunks, f, indent=2)

        return chunks_path, chunks

    # COMPLETE END-TO-END WORKFLOW (STAGE 1 -> STAGE 2 -> STAGE 3)
    def process_file(self, file_path):
        file_name = os.path.basename(file_path)
        logger.info(f"Processing File with UnifiedExcelRAGPipeline: {file_name}")

        # Stage 1: Generate Unified Markdown
        md_path, sqlite_tables = self.generate_unified_markdown(file_path)
        if not md_path:
            return [], sqlite_tables

        # Stage 2: Compress to .md.gz archive
        gz_path = self.compress_to_gz(md_path)

        # Stage 3: Stream .md.gz archive and Chunk for RAG
        chunks_path, chunks = self.chunk_compressed_archive(gz_path, file_name, file_path=file_path)
        logger.info(f"Unified Pipeline Completed: {len(chunks)} chunks & {len(sqlite_tables)} SQL tables created for {file_name}")

        return chunks, sqlite_tables
